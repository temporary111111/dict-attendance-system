"""Validated bulk import para sa official PSA PSGC Excel masterlist."""

from __future__ import annotations

from dataclasses import dataclass
from hashlib import sha256
from io import BytesIO
from numbers import Integral, Real
from pathlib import Path
import re
from typing import Any
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import PSGCBarangay, PSGCCityMunicipality, PSGCProvince, PSGCRegion
from app.services.audit_service import build_audit_log


MAX_IMPORT_ERRORS = 50
HEADER_SCAN_ROWS = 25

HEADER_ALIASES = {
    "code": {
        "psgccode",
        "psgc10digitcode",
        "10digitpsgc",
        "10digitpsgccode",
        "10digitcode",
        "geographiccode",
    },
    "name": {"areaname", "geographicname", "name", "geographicareaname"},
    "level": {
        "geographiclevel",
        "geographicallevel",
        "geoglevel",
        "level",
    },
    "region": {"regioncode", "reg"},
    "province": {"provincecode", "prov", "prv"},
    "city_municipality": {
        "citymunicipalitycode",
        "citymunicipalitypsgccode",
        "municipalitycode",
        "citycode",
        "mun",
    },
}


class PSGCImportValidationError(Exception):
    """Raised kapag hindi safe i-import ang uploaded PSGC workbook."""

    def __init__(self, errors: list[str], preview: dict[str, Any] | None = None):
        super().__init__(errors[0] if errors else "Invalid PSGC workbook.")
        self.errors = errors
        self.preview = preview or {"valid": False, "errors": errors}


@dataclass
class PSGCImportRow:
    """Isang normalized PSGC row bago ito maging database record."""

    row_number: int
    code: str
    name: str
    level: str
    city_municipality_type: str | None
    region_code: str | None
    province_code: str | None
    province_code_is_explicit: bool
    city_municipality_code: str | None


@dataclass
class ParsedPSGCWorkbook:
    """Validated rows at metadata na gamit ng preview at actual import."""

    file_name: str
    file_sha256: str
    sheet_name: str
    header_row: int
    rows: list[PSGCImportRow]


def _normalize_header(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").lower())


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, Integral) and not isinstance(value, bool):
        return str(value)
    if isinstance(value, Real) and not isinstance(value, bool):
        return str(int(value)) if value.is_integer() else str(value)
    return str(value).strip()


def _normalize_code(value: object) -> str | None:
    raw = _cell_text(value).replace(" ", "")
    if not raw:
        return None
    if raw.endswith(".0"):
        raw = raw[:-2]
    if not raw.isdigit() or len(raw) > 10:
        return None
    return raw.zfill(10)


def _normalize_level(value: object) -> tuple[str, str | None] | None:
    level = _normalize_header(value)
    if level in {"reg", "region"}:
        return "region", None
    if level in {"prov", "province"}:
        return "province", None
    if level in {"city", "highlyurbanizedcity", "componentcity", "independentcomponentcity"}:
        return "city_municipality", "city"
    if level in {"mun", "municipality"}:
        return "city_municipality", "municipality"
    if level in {"bgy", "barangay"}:
        return "barangay", None
    return None


def _revision_one_province_parent_code(code: str) -> str:
    """PSA Revision 1: region 2 digits + province/HUC 3 digits."""
    return f"{code[:5]}00000"


def _revision_one_city_parent_code(code: str) -> str:
    """PSA Revision 1: region 2 + province/HUC 3 + city/municipality 2."""
    return f"{code[:7]}000"


def _add_error(errors: list[str], message: str) -> None:
    if len(errors) < MAX_IMPORT_ERRORS:
        errors.append(message)


def _find_header_columns(sheet) -> tuple[int, dict[str, int]] | None:
    """Hinahanap ang PSA-style headers kahit may title rows bago ang data."""
    for row_number, values in enumerate(
        sheet.iter_rows(min_row=1, max_row=HEADER_SCAN_ROWS, values_only=True),
        start=1,
    ):
        normalized = [_normalize_header(value) for value in values]
        columns: dict[str, int] = {}
        for key in ("code", "name", "level", "region", "province", "city_municipality"):
            aliases = HEADER_ALIASES[key]
            for index, header in enumerate(normalized):
                if header in aliases:
                    columns[key] = index
                    break
        if {"code", "name", "level"}.issubset(columns):
            return row_number, columns
    return None


def _empty_preview(file_name: str, file_sha256: str, errors: list[str]) -> dict[str, Any]:
    return {
        "valid": False,
        "file_name": file_name,
        "file_sha256": file_sha256,
        "sheet_name": None,
        "header_row": None,
        "counts": {
            "regions": 0,
            "provinces": 0,
            "cities_municipalities": 0,
            "barangays": 0,
        },
        "errors": errors,
    }


def _build_preview(parsed: ParsedPSGCWorkbook, errors: list[str]) -> dict[str, Any]:
    counts = {
        "regions": sum(row.level == "region" for row in parsed.rows),
        "provinces": sum(row.level == "province" for row in parsed.rows),
        "cities_municipalities": sum(
            row.level == "city_municipality" for row in parsed.rows
        ),
        "barangays": sum(row.level == "barangay" for row in parsed.rows),
    }
    return {
        "valid": not errors,
        "file_name": parsed.file_name,
        "file_sha256": parsed.file_sha256,
        "sheet_name": parsed.sheet_name,
        "header_row": parsed.header_row,
        "counts": counts,
        "errors": errors,
    }


def _parent_code(
    values: tuple[object, ...],
    columns: dict[str, int],
    key: str,
    fallback: str,
) -> str | None:
    index = columns.get(key)
    if index is not None and index < len(values):
        if _cell_text(values[index]):
            return _normalize_code(values[index])
    return fallback


def _has_parent_value(
    values: tuple[object, ...],
    columns: dict[str, int],
    key: str,
) -> bool:
    index = columns.get(key)
    return index is not None and index < len(values) and bool(_cell_text(values[index]))


def _validate_hierarchy(parsed: ParsedPSGCWorkbook) -> list[str]:
    errors: list[str] = []
    region_codes = {row.code for row in parsed.rows if row.level == "region"}
    province_codes = {row.code for row in parsed.rows if row.level == "province"}
    city_codes = {
        row.code for row in parsed.rows if row.level == "city_municipality"
    }

    # Sa NCR at ibang non-province-based areas, walang valid province row.
    # Kaya blank ang local province FK kapag derived lang ang candidate code.
    for row in parsed.rows:
        if (
            row.level == "city_municipality"
            and not row.province_code_is_explicit
            and row.province_code not in province_codes
        ):
            row.province_code = None

    for row in parsed.rows:
        if row.level == "province" and row.region_code not in region_codes:
            _add_error(
                errors,
                f"Row {row.row_number}: province is missing region parent {row.region_code}.",
            )
        elif row.level == "city_municipality":
            if row.region_code not in region_codes:
                _add_error(
                    errors,
                    f"Row {row.row_number}: city or municipality is missing region parent {row.region_code}.",
                )
            elif row.province_code_is_explicit and row.province_code is None:
                _add_error(
                    errors,
                    f"Row {row.row_number}: city or municipality has an invalid province parent code.",
                )
            elif row.province_code is not None and row.province_code not in province_codes:
                _add_error(
                    errors,
                    f"Row {row.row_number}: city or municipality has unknown province parent {row.province_code}.",
                )
        elif row.level == "barangay" and row.city_municipality_code not in city_codes:
            _add_error(
                errors,
                f"Row {row.row_number}: barangay is missing city or municipality parent {row.city_municipality_code}.",
            )
    return errors


def _parse_psgc_workbook(contents: bytes, file_name: str) -> ParsedPSGCWorkbook:
    file_sha256 = sha256(contents).hexdigest()
    if Path(file_name).suffix.lower() != ".xlsx":
        raise PSGCImportValidationError(
            ["Upload an official PSGC Excel (.xlsx) file."],
            _empty_preview(file_name, file_sha256, ["Upload an official PSGC Excel (.xlsx) file."]),
        )
    if not contents:
        raise PSGCImportValidationError(
            ["The uploaded PSGC file is empty."],
            _empty_preview(file_name, file_sha256, ["The uploaded PSGC file is empty."]),
        )

    try:
        workbook = load_workbook(
            BytesIO(contents),
            read_only=True,
            data_only=True,
            keep_links=False,
        )
    except (BadZipFile, InvalidFileException, OSError, ValueError) as error:
        message = f"The uploaded file is not a readable Excel workbook: {error}"
        raise PSGCImportValidationError(
            [message],
            _empty_preview(file_name, file_sha256, [message]),
        ) from error

    selected_sheet = None
    header_row = 0
    columns: dict[str, int] = {}
    for sheet in workbook.worksheets:
        result = _find_header_columns(sheet)
        if result is not None:
            selected_sheet = sheet
            header_row, columns = result
            break
    if selected_sheet is None:
        message = "Required PSGC headers were not found: PSGC code, area name, and geographic level."
        raise PSGCImportValidationError(
            [message],
            _empty_preview(file_name, file_sha256, [message]),
        )

    errors: list[str] = []
    rows: list[PSGCImportRow] = []
    seen_codes: set[str] = set()
    current_city_municipality_code: str | None = None
    structural_group_rows: dict[str, int] = {}
    for row_number, values in enumerate(
        selected_sheet.iter_rows(min_row=header_row + 1, values_only=True),
        start=header_row + 1,
    ):
        if not any(value is not None and _cell_text(value) for value in values):
            continue
        code_value = values[columns["code"]] if columns["code"] < len(values) else None
        name_value = values[columns["name"]] if columns["name"] < len(values) else None
        level_value = values[columns["level"]] if columns["level"] < len(values) else None
        code = _normalize_code(code_value)
        name = _cell_text(name_value)
        level_data = _normalize_level(level_value)
        normalized_level = _normalize_header(level_value)

        if code is None:
            _add_error(errors, f"Row {row_number}: PSGC code must be a numeric code up to 10 digits.")
            continue
        if code in seen_codes:
            _add_error(errors, f"Row {row_number}: duplicate PSGC code {code}.")
            continue
        seen_codes.add(code)
        if not name:
            _add_error(errors, f"Row {row_number}: area name is required.")
            continue
        if len(name) > 150:
            _add_error(errors, f"Row {row_number}: area name exceeds 150 characters.")
            continue
        if normalized_level == "submun":
            if current_city_municipality_code is None:
                _add_error(
                    errors,
                    f"Row {row_number}: submunicipality is missing its parent city or municipality.",
                )
            continue
        if not normalized_level:
            # May dalawang PSA grouping rows na walang geographic level. Hindi sila
            # lookup level sa schema, pero sinusundan sila ng actual city/municipality.
            structural_group_rows[code] = row_number
            current_city_municipality_code = None
            continue
        if level_data is None:
            _add_error(errors, f"Row {row_number}: unsupported geographic level '{_cell_text(level_value)}'.")
            continue

        level, city_municipality_type = level_data
        if level in {"region", "province"}:
            current_city_municipality_code = None
        if level == "city_municipality":
            structural_group_rows.pop(f"{code[:6]}0000", None)
            current_city_municipality_code = code
        rows.append(
            PSGCImportRow(
                row_number=row_number,
                code=code,
                name=name,
                level=level,
                city_municipality_type=city_municipality_type,
                region_code=(
                    None
                    if level == "region"
                    else _parent_code(values, columns, "region", f"{code[:2]}00000000")
                ),
                province_code=(
                    _parent_code(
                        values,
                        columns,
                        "province",
                        _revision_one_province_parent_code(code),
                    )
                    if level == "city_municipality"
                    else None
                ),
                province_code_is_explicit=(
                    level == "city_municipality"
                    and _has_parent_value(values, columns, "province")
                ),
                city_municipality_code=(
                    (
                        _parent_code(
                            values,
                            columns,
                            "city_municipality",
                            _revision_one_city_parent_code(code),
                        )
                        if _has_parent_value(values, columns, "city_municipality")
                        else current_city_municipality_code
                        or _revision_one_city_parent_code(code)
                    )
                    if level == "barangay"
                    else None
                ),
            )
        )

    parsed = ParsedPSGCWorkbook(
        file_name=file_name,
        file_sha256=file_sha256,
        sheet_name=selected_sheet.title,
        header_row=header_row,
        rows=rows,
    )
    for code, row_number in structural_group_rows.items():
        _add_error(
            errors,
            f"Row {row_number}: structural PSGC group {code} is not followed by a city or municipality.",
        )
    errors.extend(_validate_hierarchy(parsed))
    if not rows:
        _add_error(errors, "No PSGC data rows were found in the selected worksheet.")
    preview = _build_preview(parsed, errors)
    if errors:
        raise PSGCImportValidationError(errors, preview)
    return parsed


def preview_psgc_workbook(contents: bytes, file_name: str) -> dict[str, Any]:
    """Validates the Excel file first; this function never writes to MySQL."""
    parsed = _parse_psgc_workbook(contents, file_name)
    return _build_preview(parsed, [])


def _upsert_counts(
    record: object | None,
    values: dict[str, Any],
    created: dict[str, int],
    updated: dict[str, int],
    unchanged: dict[str, int],
    level: str,
) -> object:
    if record is None:
        created[level] += 1
        return None
    if all(getattr(record, key) == value for key, value in values.items()):
        unchanged[level] += 1
    else:
        updated[level] += 1
    return record


def import_psgc_workbook(
    db: Session,
    contents: bytes,
    file_name: str,
    source_version: str,
    user_id: int,
) -> dict[str, Any]:
    """Imports validated rows in parent-to-child order inside one transaction."""
    parsed = _parse_psgc_workbook(contents, file_name)
    current_regions = {row.region_code: row for row in db.scalars(select(PSGCRegion)).all()}
    current_provinces = {row.province_code: row for row in db.scalars(select(PSGCProvince)).all()}
    current_cities = {
        row.city_municipality_code: row
        for row in db.scalars(select(PSGCCityMunicipality)).all()
    }
    current_barangays = {row.barangay_code: row for row in db.scalars(select(PSGCBarangay)).all()}
    imported_province_codes = {
        row.code for row in parsed.rows if row.level == "province"
    }
    created = dict.fromkeys(("regions", "provinces", "cities_municipalities", "barangays"), 0)
    updated = dict.fromkeys(("regions", "provinces", "cities_municipalities", "barangays"), 0)
    unchanged = dict.fromkeys(("regions", "provinces", "cities_municipalities", "barangays"), 0)

    for row in parsed.rows:
        if row.level == "region":
            values = {"region_name": row.name, "is_active": True}
            record = _upsert_counts(current_regions.get(row.code), values, created, updated, unchanged, "regions")
            if record is None:
                record = PSGCRegion(region_code=row.code, **values)
            else:
                for key, value in values.items():
                    setattr(record, key, value)
            db.add(record)
        elif row.level == "province":
            values = {"region_code": row.region_code, "province_name": row.name, "is_active": True}
            record = _upsert_counts(current_provinces.get(row.code), values, created, updated, unchanged, "provinces")
            if record is None:
                record = PSGCProvince(province_code=row.code, **values)
            else:
                for key, value in values.items():
                    setattr(record, key, value)
            db.add(record)
        elif row.level == "city_municipality":
            values = {
                "region_code": row.region_code,
                "province_code": (
                    row.province_code
                    if row.province_code in imported_province_codes
                    else None
                ),
                "city_municipality_name": row.name,
                "city_municipality_type": row.city_municipality_type,
                "is_active": True,
            }
            record = _upsert_counts(current_cities.get(row.code), values, created, updated, unchanged, "cities_municipalities")
            if record is None:
                record = PSGCCityMunicipality(city_municipality_code=row.code, **values)
            else:
                for key, value in values.items():
                    setattr(record, key, value)
            db.add(record)
        elif row.level == "barangay":
            values = {
                "city_municipality_code": row.city_municipality_code,
                "barangay_name": row.name,
                "is_active": True,
            }
            record = _upsert_counts(current_barangays.get(row.code), values, created, updated, unchanged, "barangays")
            if record is None:
                record = PSGCBarangay(barangay_code=row.code, **values)
            else:
                for key, value in values.items():
                    setattr(record, key, value)
            db.add(record)

    result = {
        "source_version": source_version,
        "file_name": parsed.file_name,
        "file_sha256": parsed.file_sha256,
        "counts": _build_preview(parsed, [])["counts"],
        "created": created,
        "updated": updated,
        "unchanged": unchanged,
    }
    db.add(
        build_audit_log(
            user_id=user_id,
            action="imported_psgc_masterlist",
            entity_type="psgc_import",
            entity_id=None,
            description=f"Imported PSGC masterlist {source_version}.",
            old_values=None,
            new_values=result,
            ip_address=None,
            user_agent=None,
        )
    )
    try:
        db.commit()
    except Exception:
        db.rollback()
        raise
    return result
